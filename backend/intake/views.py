"""
API views for the Gate-Intake platform.

Operational ownership sits with the **Contractor**: they own their worker pool,
run the split-pane intake workbench, administer trade tests, track safety-video
progress, bulk-import workers, scan resumes and search their own labour pool by
skill demand. The Principal Employer reviews submitted lists and nothing else.
Gate Security scans an Aadhaar and gets a real-time GREEN/RED decision.

Endpoints (all under /api/):
  POST   auth/login/                          -> token + persona
  GET    me/                                  -> current user
  GET    requirements/                        -> master requirement catalogue
  GET    projects/                            -> list projects (role scoped, read-only)
  GET    projects/<id>/eligible-workers/      -> compliance split for a contractor
  POST   workforce-demand/                    -> "3 Carpenters, 4 Masons" pool search
  GET    workers/  POST workers/              -> Contractor's own worker registry
  DELETE workers/<id>/                        -> Contractor removes their worker
  GET    verification-status/                 -> whole-pool verification matrix
  POST   workers/bulk-upload/                 -> Contractor CSV/Excel import
  POST   documents/upload/                    -> inline gap-fix upload
  PATCH  documents/<id>/review/               -> PE verify/reject a document
  GET/POST intake-lists/                      -> list / submit
  PATCH  intake-lists/<id>/review/            -> PE approve / request changes / reject
  GET    gate-check/?aadhar=<n>               -> REAL-TIME gate decision
  POST   intake/onboard-worker/               -> unified 6-document single-pass intake
  POST   intake/verify-document/              -> commit one verified document
  POST   intake/ocr-extract/                  -> OCR a scan into form fields
  POST   resume/parse/                        -> parse a resume without committing
  GET    candidates/search/                   -> multi-attribute candidate filter
  POST   storage/signed-url/                  -> batch of fresh expiring links
  GET    trade-test/start/  POST submit-attempt/
  POST   safety-video/heartbeat/
"""
from __future__ import annotations

import csv
import io
from datetime import timedelta

from django.conf import settings
from django.db import IntegrityError, transaction
from django.http import FileResponse, Http404
from django.shortcuts import get_object_or_404
from django.utils import timezone
from rest_framework import status
from rest_framework.authtoken.models import Token
from rest_framework.decorators import api_view, permission_classes
from rest_framework.parsers import FormParser, JSONParser, MultiPartParser
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from . import crypto, doctype, ocr, resume as resume_parser, storage
from .models import (
    INTAKE_EXPIRY_DAYS,
    CandidateProfile,
    CandidateSkill,
    IntakeList,
    IntakeListWorker,
    IntakeMedicalRecord,
    IntakePoliceVerification,
    Project,
    RequirementMaster,
    SafetyTrainingProgress,
    Skill,
    TradeTestAttempt,
    TradeTestQuestion,
    User,
    Worker,
    WorkerBankAccount,
    WorkerDocument,
    category_for_skill,
)
from .serializers import (
    CandidateProfileSerializer,
    WorkerBankAccountSerializer,
    IntakeListSerializer,
    IntakeMedicalRecordSerializer,
    IntakePoliceVerificationSerializer,
    ProjectSerializer,
    RequirementMasterSerializer,
    TradeTestQuestionSerializer,
    UserSerializer,
    WorkerDocumentSerializer,
    WorkerSerializer,
)

TRADE_TEST_QUESTION_COUNT = 5
TRADE_TEST_PASS_MARK = 3
TRADE_TEST_MAX_ATTEMPTS = 3


# ---------------------------------------------------------------------------
# Small role / scope helpers
# ---------------------------------------------------------------------------
def _require_role(request, *roles) -> Response | None:
    """Return a 403 Response if the user's role is not in ``roles``; else None."""
    if request.user.role not in roles:
        return Response(
            {
                "detail": f"This action requires role(s): {', '.join(roles)}. "
                f"You are '{request.user.role}'."
            },
            status=status.HTTP_403_FORBIDDEN,
        )
    return None


def _worker_scope(request):
    """Workers this user may see.

    A contractor sees only their own pool — that scoping is the whole point of
    moving operations to them. PE and Gate see everything (read paths only).
    """
    if request.user.role == User.Role.CONTRACTOR:
        return Worker.objects.filter(contractor=request.user)
    return Worker.objects.all()


def _prefetched(queryset):
    """Prefetch every relation the compliance engine touches (kills the N+1)."""
    return queryset.select_related("safety_video", "candidate_profile").prefetch_related(
        "documents__requirement",
        "medical_records",
        "police_verifications",
        "trade_test_attempts",
        "candidate_profile__candidate_skills__skill",
        "bank_account",
    )


def _owned_worker(request, pk) -> Worker:
    """Fetch a worker the caller is allowed to act on, else 404."""
    return get_object_or_404(_worker_scope(request), pk=pk)


def _serializer_context(request, *, sign=False):
    """PII is revealed to the owning contractor and to the reviewing PE only."""
    return {
        "request": request,
        "sign": sign,
        "reveal_pii": request.user.role in (
            User.Role.CONTRACTOR,
            User.Role.PRINCIPAL_EMPLOYER,
        ),
    }


def _store_upload(upload, prefix: str) -> str | None:
    """Push one uploaded file to the private bucket, return its object key."""
    if upload is None:
        return None
    data = upload.read()
    if not data:
        return None
    return storage.upload(data, upload.content_type, prefix, upload.name)


def _iso(d):
    return d.isoformat() if d else None


def _parse_iso_date(value):
    """Lenient ISO date parse — returns a ``date`` or None."""
    if not value:
        return None
    try:
        return timezone.datetime.fromisoformat(str(value)).date()
    except ValueError:
        return None


def _slot_for(doc_type: str, requirement_name: str = "") -> str:
    """Map a doc_type (+ requirement) onto the intake slot it belongs to."""
    if doc_type == "IDENTITY":
        return {
            "Aadhar": "aadhaar",
            "PAN": "pan",
            "Safety Training": "safety",
        }.get(requirement_name, "aadhaar")
    return {"MEDICAL": "medical", "POLICE": "pvc", "BANK": "bank"}.get(doc_type, doc_type.lower())


_SLOT_DOC_TYPE = {
    "aadhaar": ("IDENTITY", "Aadhar"),
    "pan": ("IDENTITY", "PAN"),
    "safety": ("IDENTITY", "Safety Training"),
    "medical": ("MEDICAL", ""),
    "pvc": ("POLICE", ""),
    "bank": ("BANK", ""),
}


def _doc_type_for_slot(slot: str) -> str:
    return _SLOT_DOC_TYPE.get(slot, ("IDENTITY", ""))[0]


def _requirement_for_slot(slot: str) -> str:
    return _SLOT_DOC_TYPE.get(slot, ("IDENTITY", ""))[1]


def _as_bool(value):
    """Multipart sends booleans as strings ("true"/"false"); coerce safely."""
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"1", "true", "yes", "on"}


# ---------------------------------------------------------------------------
# Auth
# ---------------------------------------------------------------------------
class LoginView(APIView):
    """Email/password login returning a DRF token and the persona payload.

    The frontend role-switcher calls this with each hardcoded credential to
    obtain that persona's token, then masquerades by swapping the active token.
    """

    permission_classes = [AllowAny]

    def post(self, request):
        email = (request.data.get("email") or "").strip().lower()
        password = request.data.get("password") or ""

        user = User.objects.filter(email=email).first()
        if user is None or not user.check_password(password):
            return Response(
                {"detail": "Invalid credentials."},
                status=status.HTTP_401_UNAUTHORIZED,
            )

        token, _ = Token.objects.get_or_create(user=user)
        return Response({"token": token.key, "user": UserSerializer(user).data})


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def me(request):
    return Response(UserSerializer(request.user).data)


# ---------------------------------------------------------------------------
# Requirements catalogue
# ---------------------------------------------------------------------------
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def requirements(request):
    """GET /api/requirements/ — the full master requirement catalogue.

    Powers the contractor's checkbox filter (search workers by which
    requirements they have fulfilled, independent of any project's mandatory set).
    """
    qs = RequirementMaster.objects.all()
    return Response(RequirementMasterSerializer(qs, many=True).data)


# ---------------------------------------------------------------------------
# Projects — read-only
#
# Project and requirement configuration was removed from the product: the PE
# dashboard is strictly a review surface now, so there is no create/update path
# here. Projects arrive via the seed (or an admin) and are consumed read-only.
# ---------------------------------------------------------------------------
class ProjectListView(APIView):
    def get(self, request):
        user = request.user
        if user.role == User.Role.PRINCIPAL_EMPLOYER:
            qs = Project.objects.filter(principal_employer=user)
        elif user.role == User.Role.CONTRACTOR:
            qs = Project.objects.filter(contractors=user)
        else:
            qs = Project.objects.all()
        qs = qs.prefetch_related("project_requirements__requirement", "contractors")
        return Response(ProjectSerializer(qs, many=True).data)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def project_detail(request, pk):
    project = get_object_or_404(
        Project.objects.prefetch_related("project_requirements__requirement"), pk=pk
    )
    return Response(ProjectSerializer(project).data)


class EligibleWorkersView(APIView):
    """
    GET /api/projects/<id>/eligible-workers/

    Evaluates the project's mandatory requirements against the calling
    contractor's own workers and returns a structured split:

      {
        "project": {...},
        "required_documents": [...],
        "summary": {"total": 8, "ready": 5, "needs_fixes": 3},
        "ready_to_deploy": [ {worker + compliance}, ... ],
        "needs_fixes":     [ {worker + compliance (with explicit gaps)}, ... ]
      }
    """

    def get(self, request, pk):
        project = get_object_or_404(
            Project.objects.prefetch_related("project_requirements__requirement"),
            pk=pk,
        )

        workers_qs = _worker_scope(request)
        if request.user.role != User.Role.CONTRACTOR:
            contractor_id = request.query_params.get("contractor_id")
            if contractor_id:
                workers_qs = workers_qs.filter(contractor_id=contractor_id)
        workers_qs = _prefetched(workers_qs)

        required = [
            {
                "requirement_id": pr.requirement.id,
                "requirement_name": pr.requirement.name,
                "is_expirable": pr.requirement.is_expirable,
            }
            for pr in project.project_requirements.filter(is_mandatory=True)
        ]

        context = _serializer_context(request)
        ready, needs_fixes = [], []
        for worker in workers_qs:
            compliance = worker.compliance_against_project(project)
            payload = {
                "worker": WorkerSerializer(worker, context=context).data,
                "compliance": compliance,
            }
            (ready if compliance["is_compliant"] else needs_fixes).append(payload)

        return Response(
            {
                "project": ProjectSerializer(project).data,
                "required_documents": required,
                "summary": {
                    "total": len(ready) + len(needs_fixes),
                    "ready": len(ready),
                    "needs_fixes": len(needs_fixes),
                },
                "ready_to_deploy": ready,
                "needs_fixes": needs_fixes,
            }
        )


# ---------------------------------------------------------------------------
# Workforce demand — "3 Carpenters, 4 Masons" against the contractor's own pool
# ---------------------------------------------------------------------------
class WorkforceDemandView(APIView):
    """
    POST /api/workforce-demand/   (Contractor)

    Body::

        {"demands": [{"skill": "Carpenter", "count": 3},
                     {"skill": "Mason", "count": 4}],
         "project": 12}          # optional — scopes document requirements

    The contractor types their immediate needs; this instantly queries their own
    labour pool and answers, per line: who is deployable right now, who could be
    with fixes, and how many bodies short they are.

    Matching is deliberately broad — the worker's ``skill_type`` OR any skill
    parsed from their resume — because "Carpenter" on a CV and "carpentry" in
    the registry are the same person.
    """

    def post(self, request):
        denied = _require_role(request, User.Role.CONTRACTOR)
        if denied:
            return denied

        raw_demands = request.data.get("demands") or []
        if not isinstance(raw_demands, list) or not raw_demands:
            return Response(
                {"detail": "demands must be a non-empty list of {skill, count}."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        project = None
        project_id = request.data.get("project")
        if project_id:
            project = get_object_or_404(
                Project.objects.prefetch_related("project_requirements__requirement"),
                pk=project_id,
            )

        workers = list(_prefetched(_worker_scope(request)))

        # Evaluate each worker once — several demand lines usually overlap.
        evaluated = {w.id: w.compliance_snapshot(project) for w in workers}
        context = _serializer_context(request)

        def matches(worker, skill: str) -> bool:
            needle = skill.strip().lower()
            if not needle:
                return False
            if needle in (worker.skill_type or "").lower():
                return True
            try:
                profile = worker.candidate_profile
            except CandidateProfile.DoesNotExist:
                return False
            return any(
                needle in cs.skill.name for cs in profile.candidate_skills.all()
            )

        lines, total_required, total_ready = [], 0, 0
        for raw in raw_demands:
            skill = str((raw or {}).get("skill") or "").strip()
            try:
                count = max(0, int((raw or {}).get("count") or 0))
            except (TypeError, ValueError):
                count = 0
            if not skill:
                continue

            matched = [w for w in workers if matches(w, skill)]
            ready = [w for w in matched if evaluated[w.id]["is_compliant"]]
            fixable = [w for w in matched if not evaluated[w.id]["is_compliant"]]

            total_required += count
            total_ready += min(count, len(ready))

            lines.append(
                {
                    "skill": skill,
                    "required": count,
                    "available": len(ready),
                    "shortfall": max(0, count - len(ready)),
                    # A shortfall the contractor can close today by fixing papers
                    # rather than by hiring — the actionable number.
                    "fixable": len(fixable),
                    "ready_workers": [
                        {
                            "worker": WorkerSerializer(w, context=context).data,
                            "compliance": evaluated[w.id],
                        }
                        for w in ready[: max(count, 10)]
                    ],
                    "needs_fixes": [
                        {
                            "worker": WorkerSerializer(w, context=context).data,
                            "compliance": evaluated[w.id],
                        }
                        for w in fixable[:10]
                    ],
                }
            )

        return Response(
            {
                "project": ProjectSerializer(project).data if project else None,
                "summary": {
                    "total_required": total_required,
                    "total_ready": total_ready,
                    "total_shortfall": max(0, total_required - total_ready),
                    "pool_size": len(workers),
                },
                "lines": lines,
            }
        )


# ---------------------------------------------------------------------------
# Workers — bulk import
# ---------------------------------------------------------------------------
class WorkerBulkUploadView(APIView):
    """
    POST /api/workers/bulk-upload/   (Contractor)

    Accepts a CSV (or Excel .xlsx) file under the form field ``file`` with
    columns: name, aadhar_number, skill_type.

    Imported workers are always assigned to the calling contractor — the pool is
    theirs, so there is no contractor column to spoof.

    Idempotent-ish: existing Aadhar numbers are reported as skipped, never
    duplicated, honouring the UNIQUE constraint.
    """

    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        denied = _require_role(request, User.Role.CONTRACTOR)
        if denied:
            return denied

        upload = request.FILES.get("file")
        if upload is None:
            return Response(
                {"detail": "No file provided under form field 'file'."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            rows = self._parse_rows(upload)
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        created, skipped, errors = [], [], []

        for idx, row in enumerate(rows, start=2):  # row 1 is the header
            name = (row.get("name") or "").strip()
            aadhar = str(row.get("aadhar_number") or "").strip()
            skill = (row.get("skill_type") or "").strip()

            if not (name and aadhar and skill):
                errors.append({"row": idx, "error": "Missing name/aadhar/skill."})
                continue
            if len(aadhar) != 12 or not aadhar.isdigit():
                errors.append({"row": idx, "error": f"Invalid Aadhar '{aadhar}'."})
                continue

            try:
                with transaction.atomic():
                    worker = Worker.objects.create(
                        name=name,
                        aadhar_number=aadhar,
                        skill_type=skill,
                        contractor=request.user,
                    )
                created.append({"row": idx, "id": worker.id, "aadhar": aadhar})
            except IntegrityError:
                skipped.append({"row": idx, "aadhar": aadhar, "reason": "duplicate"})

        return Response(
            {
                "created_count": len(created),
                "skipped_count": len(skipped),
                "error_count": len(errors),
                "created": created,
                "skipped": skipped,
                "errors": errors,
            },
            status=status.HTTP_201_CREATED if created else status.HTTP_200_OK,
        )

    @staticmethod
    def _parse_rows(upload) -> list[dict]:
        """Return a list of dict rows from a CSV or XLSX upload."""
        filename = (upload.name or "").lower()

        if filename.endswith(".xlsx"):
            try:
                from openpyxl import load_workbook
            except ImportError as exc:  # pragma: no cover
                raise ValueError("openpyxl not installed for .xlsx parsing.") from exc

            wb = load_workbook(upload, read_only=True, data_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header = [str(h).strip().lower() if h else "" for h in next(rows_iter)]
            except StopIteration:
                return []
            result = []
            for values in rows_iter:
                result.append(
                    {header[i]: values[i] for i in range(len(header)) if i < len(values)}
                )
            return result

        # Default: CSV
        decoded = upload.read().decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(decoded))
        normalized = []
        for raw in reader:
            normalized.append({(k or "").strip().lower(): v for k, v in raw.items()})
        return normalized


class WorkerListView(APIView):
    """
    GET  /api/workers/  — role-scoped worker registry (used by dashboards).
    POST /api/workers/  — Contractor creates a single worker from scratch.
    """

    def get(self, request):
        qs = _prefetched(_worker_scope(request))
        return Response(
            WorkerSerializer(qs, many=True, context=_serializer_context(request)).data
        )

    def post(self, request):
        denied = _require_role(request, User.Role.CONTRACTOR)
        if denied:
            return denied

        name = (request.data.get("name") or "").strip()
        aadhar = (request.data.get("aadhar_number") or "").strip()
        skill = (request.data.get("skill_type") or "").strip()

        if not (name and aadhar and skill):
            return Response(
                {"detail": "name, aadhar_number and skill_type are required."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if len(aadhar) != 12 or not aadhar.isdigit():
            return Response(
                {"detail": "Aadhar number must be exactly 12 digits."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            worker = Worker.objects.create(
                name=name,
                aadhar_number=aadhar,
                skill_type=skill,
                contractor=request.user,
            )
        except IntegrityError:
            return Response(
                {"detail": f"A worker with Aadhar {aadhar} already exists."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        return Response(
            WorkerSerializer(worker, context=_serializer_context(request)).data,
            status=status.HTTP_201_CREATED,
        )


class WorkerDetailView(APIView):
    """DELETE /api/workers/<id>/ — Contractor removes one of their own workers
    and all their records (documents, medical, police, trade-test attempts,
    safety video, candidate profile and deployment-list memberships cascade)."""

    def delete(self, request, pk):
        denied = _require_role(request, User.Role.CONTRACTOR)
        if denied:
            return denied
        worker = _owned_worker(request, pk)

        # Clean the private bucket up as well, so deleting a worker really does
        # remove their documents rather than orphaning them.
        keys = [d.storage_key for d in worker.documents.all() if d.storage_key]
        keys += [m.storage_key for m in worker.medical_records.all() if m.storage_key]
        keys += [p.storage_key for p in worker.police_verifications.all() if p.storage_key]
        try:
            keys.append(worker.candidate_profile.resume_key)
        except CandidateProfile.DoesNotExist:
            pass

        name = worker.name
        worker.delete()
        for key in keys:
            storage.delete(key)
        return Response({"deleted": True, "name": name})


class VerificationStatusView(APIView):
    """
    GET /api/verification-status/   (Contractor)

    A whole-pool verification matrix: one row per worker with the status of
    every verification type and a fresh expiring link to each stored document,
    so the contractor sees at a glance what is verified vs remaining and can
    re-open or download any scan.
    """

    # status values grouped into "done" vs "remaining" for the summary count.
    _DONE = {"VERIFIED", "PASSED"}

    def get(self, request):
        denied = _require_role(request, User.Role.CONTRACTOR)
        if denied:
            return denied

        today = timezone.now().date()
        reqs = {r.name: r for r in RequirementMaster.objects.all()}
        workers = _prefetched(_worker_scope(request)).order_by("name")

        def link(obj):
            if obj is None:
                return None
            if getattr(obj, "storage_key", ""):
                return storage.signed_url(obj.storage_key)
            legacy = getattr(obj, "document_file", None)
            if legacy:
                try:
                    return request.build_absolute_uri(legacy.url)
                except ValueError:
                    pass
            return getattr(obj, "file_url", "") or None

        rows = []
        for w in workers:
            # Most-recent document per requirement.
            latest_doc = {}
            for d in sorted(w.documents.all(), key=lambda x: x.uploaded_at):
                latest_doc[d.requirement_id] = d

            items = []

            # --- Document requirements: Aadhaar, PAN, Safety Cert ---
            for name, label in (("Aadhar", "Aadhaar"), ("PAN", "PAN"),
                                ("Safety Training", "Safety Cert")):
                req = reqs.get(name)
                d = latest_doc.get(req.id) if req else None
                if d is None:
                    items.append({"key": name, "label": label,
                                  "status": "MISSING", "doc_url": None})
                    continue
                st = d.verification_status.upper()  # VERIFIED / PENDING / REJECTED
                if (st == "VERIFIED" and req.is_expirable and d.expiry_date
                        and d.expiry_date < today):
                    st = "EXPIRED"
                items.append({"key": name, "label": label, "status": st,
                              "doc_url": link(d)})

            # --- Medical ---
            med = max(w.medical_records.all(), key=lambda m: m.exam_date, default=None)
            if med is None:
                m_st = "MISSING"
            elif med.expiry_date and med.expiry_date < today:
                m_st = "EXPIRED"
            elif med.color_blindness or med.vertigo:
                m_st = "FAILED"
            else:
                m_st = "VERIFIED"
            items.append({"key": "MEDICAL", "label": "Medical", "status": m_st,
                          "doc_url": link(med)})

            # --- Police verification ---
            pol = max(w.police_verifications.all(), key=lambda p: p.issue_date, default=None)
            if pol is None:
                p_st = "MISSING"
            elif pol.verification_status != WorkerDocument.Status.VERIFIED:
                p_st = "PENDING"
            elif pol.expiry_date and pol.expiry_date < today:
                p_st = "EXPIRED"
            else:
                p_st = "VERIFIED"
            items.append({"key": "POLICE", "label": "Police", "status": p_st,
                          "doc_url": link(pol)})

            # --- Trade test (no document) ---
            items.append({"key": "TRADE_TEST", "label": "Trade Test",
                          "status": w.trade_test_status, "doc_url": None})

            # --- Safety video (no document) ---
            try:
                sv = w.safety_video
            except SafetyTrainingProgress.DoesNotExist:
                sv = None
            items.append({"key": "SAFETY_VIDEO", "label": "Safety Video",
                          "status": "VERIFIED" if (sv and sv.is_completed) else "INCOMPLETE",
                          "doc_url": None})

            # --- Resume ---
            try:
                profile = w.candidate_profile
            except CandidateProfile.DoesNotExist:
                profile = None
            items.append({
                "key": "RESUME",
                "label": "Resume",
                "status": "VERIFIED" if (profile and profile.resume_key) else "MISSING",
                "doc_url": storage.signed_url(profile.resume_key) if profile else None,
            })

            # --- Bank account (payroll, not a gate pillar) ---
            try:
                bank = w.bank_account
            except WorkerBankAccount.DoesNotExist:
                bank = None
            items.append({
                "key": "BANK",
                "label": "Bank",
                "status": "VERIFIED" if (bank and bank.account_number_encrypted) else "MISSING",
                "doc_url": link(bank) if bank else None,
            })

            remaining = sum(1 for it in items if it["status"] not in self._DONE)
            rows.append({
                "id": w.id,
                "name": w.name,
                "skill_type": w.skill_type,
                "aadhar_number": w.aadhar_number,
                "items": items,
                "remaining": remaining,
                "all_verified": remaining == 0,
            })

        return Response(rows)


# ---------------------------------------------------------------------------
# Documents — inline gap-fix upload
# ---------------------------------------------------------------------------
class DocumentUploadView(APIView):
    """
    POST /api/documents/upload/   (Contractor)

    Creates or updates the document for a (worker, requirement) slot and pushes
    the file to the private bucket. Uploading a fresh document resets its status
    to 'Pending' and clears any prior rejection.

    Form fields: worker, requirement, document_number, expiry_date, file
    """

    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def post(self, request):
        denied = _require_role(request, User.Role.CONTRACTOR)
        if denied:
            return denied

        worker_id = request.data.get("worker")
        requirement_id = request.data.get("requirement")
        if not worker_id or not requirement_id:
            return Response(
                {"detail": "'worker' and 'requirement' are required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        worker = _owned_worker(request, worker_id)
        requirement = get_object_or_404(RequirementMaster, pk=requirement_id)

        # One document slot per (worker, requirement): update in place if present.
        doc, _created = WorkerDocument.objects.get_or_create(
            worker=worker, requirement=requirement
        )
        doc.document_number = request.data.get("document_number", doc.document_number)
        doc.expiry_date = request.data.get("expiry_date") or doc.expiry_date

        if "file" in request.FILES:
            old_key = doc.storage_key
            doc.storage_key = _store_upload(request.FILES["file"], "worker_docs") or ""
            doc.file_url = ""
            if old_key and old_key != doc.storage_key:
                storage.delete(old_key)
        elif request.data.get("file_url"):
            doc.file_url = request.data["file_url"]

        # A re-upload always re-enters the verification queue.
        doc.verification_status = WorkerDocument.Status.PENDING
        doc.rejection_reason = ""
        doc.save()

        return Response(
            WorkerDocumentSerializer(
                doc, context=_serializer_context(request, sign=True)
            ).data,
            status=status.HTTP_200_OK,
        )


class DocumentReviewView(APIView):
    """
    PATCH /api/documents/<id>/review/   (PE)

    Body: {"verification_status": "Verified"|"Rejected", "rejection_reason": "..."}
    """

    def patch(self, request, pk):
        denied = _require_role(request, User.Role.PRINCIPAL_EMPLOYER)
        if denied:
            return denied
        doc = get_object_or_404(WorkerDocument, pk=pk)
        new_status = request.data.get("verification_status")
        if new_status not in WorkerDocument.Status.values:
            return Response(
                {"detail": f"Invalid status. Use one of {WorkerDocument.Status.values}."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        doc.verification_status = new_status
        doc.rejection_reason = (
            request.data.get("rejection_reason", "")
            if new_status == WorkerDocument.Status.REJECTED
            else ""
        )
        doc.save()
        return Response(
            WorkerDocumentSerializer(doc, context=_serializer_context(request)).data
        )


# ---------------------------------------------------------------------------
# Intake lists — submit & review
# ---------------------------------------------------------------------------
class IntakeListView(APIView):
    """
    GET  /api/intake-lists/            -> role-scoped lists
    POST /api/intake-lists/            -> Contractor submits a finalized list

    POST body:
      {"project": <id>, "worker_ids": [1,2,3], "submit": true}
    When submit is true the list goes straight to 'Submitted'; otherwise 'Draft'.
    Only fully-compliant workers are accepted onto a submitted list.
    """

    def get(self, request):
        user = request.user
        if user.role == User.Role.CONTRACTOR:
            qs = IntakeList.objects.filter(contractor=user)
        elif user.role == User.Role.PRINCIPAL_EMPLOYER:
            qs = IntakeList.objects.filter(project__principal_employer=user)
        else:
            qs = IntakeList.objects.all()
        qs = qs.select_related("project", "contractor").prefetch_related(
            "list_workers__worker__documents__requirement",
            "list_workers__worker__candidate_profile__candidate_skills__skill",
        )
        return Response(
            IntakeListSerializer(
                qs, many=True, context=_serializer_context(request, sign=True)
            ).data
        )

    def post(self, request):
        denied = _require_role(request, User.Role.CONTRACTOR)
        if denied:
            return denied

        project = get_object_or_404(Project, pk=request.data.get("project"))
        worker_ids = request.data.get("worker_ids") or []
        submit = bool(request.data.get("submit", True))

        if not worker_ids:
            return Response(
                {"detail": "worker_ids must contain at least one worker."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        workers = list(
            _prefetched(Worker.objects.filter(id__in=worker_ids, contractor=request.user))
        )
        if len(workers) != len(set(worker_ids)):
            return Response(
                {"detail": "Some workers were not found or are not yours."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Guard: on submission every worker must be fully compliant.
        if submit:
            non_compliant = [
                w.name
                for w in workers
                if not w.compliance_against_project(project)["is_compliant"]
            ]
            if non_compliant:
                return Response(
                    {
                        "detail": "Cannot submit: these workers are not compliant.",
                        "non_compliant_workers": non_compliant,
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

        with transaction.atomic():
            intake_list = IntakeList.objects.create(
                project=project,
                contractor=request.user,
                status=IntakeList.Status.SUBMITTED if submit else IntakeList.Status.DRAFT,
                submitted_at=timezone.now() if submit else None,
            )
            IntakeListWorker.objects.bulk_create(
                [IntakeListWorker(intake_list=intake_list, worker=w) for w in workers]
            )

        return Response(
            IntakeListSerializer(intake_list, context=_serializer_context(request)).data,
            status=status.HTTP_201_CREATED,
        )


class IntakeListDetailView(APIView):
    """
    Contractor-owned edit / resubmit of a single list (the revise-in-place loop).

    GET   /api/intake-lists/<id>/           -> detail
    PATCH /api/intake-lists/<id>/           -> edit roster and/or resubmit

    PATCH body (all optional):
      {"worker_ids": [1,2], "submit": true}
    Only 'Draft' or 'Revision_Requested' lists may be edited. Omitting
    worker_ids keeps the existing roster (typical after fixing documents).
    Resubmitting flips the SAME list back to 'Submitted' and clears the prior
    PE verdict so it re-enters review as the same list id.
    """

    EDITABLE = {IntakeList.Status.DRAFT, IntakeList.Status.REVISION_REQUESTED}

    def get(self, request, pk):
        intake_list = get_object_or_404(
            IntakeList.objects.select_related("project", "contractor").prefetch_related(
                "list_workers__worker__documents__requirement"
            ),
            pk=pk,
        )
        return Response(
            IntakeListSerializer(
                intake_list, context=_serializer_context(request, sign=True)
            ).data
        )

    def patch(self, request, pk):
        denied = _require_role(request, User.Role.CONTRACTOR)
        if denied:
            return denied

        intake_list = get_object_or_404(
            IntakeList.objects.select_related("project"), pk=pk
        )
        if intake_list.contractor_id != request.user.id:
            return Response(
                {"detail": "You can only edit your own lists."},
                status=status.HTTP_403_FORBIDDEN,
            )
        if intake_list.status not in self.EDITABLE:
            return Response(
                {
                    "detail": f"Only Draft or Revision_Requested lists can be edited "
                    f"(this list is '{intake_list.status}')."
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        project = intake_list.project
        resubmit = bool(request.data.get("submit", True))

        # Roster: use provided worker_ids, else keep the existing membership.
        if "worker_ids" in request.data:
            worker_ids = request.data.get("worker_ids") or []
            if not worker_ids:
                return Response(
                    {"detail": "worker_ids must contain at least one worker."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            workers = list(
                _prefetched(
                    Worker.objects.filter(id__in=worker_ids, contractor=request.user)
                )
            )
            if len(workers) != len(set(worker_ids)):
                return Response(
                    {"detail": "Some workers were not found or are not yours."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        else:
            workers = [ilw.worker for ilw in intake_list.list_workers.all()]

        # On resubmission every worker must (again) be fully compliant.
        if resubmit:
            non_compliant = [
                w.name
                for w in workers
                if not w.compliance_against_project(project)["is_compliant"]
            ]
            if non_compliant:
                return Response(
                    {
                        "detail": "Cannot resubmit: these workers are still not compliant.",
                        "non_compliant_workers": non_compliant,
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

        with transaction.atomic():
            if "worker_ids" in request.data:
                intake_list.list_workers.all().delete()
                IntakeListWorker.objects.bulk_create(
                    [IntakeListWorker(intake_list=intake_list, worker=w) for w in workers]
                )
            if resubmit:
                intake_list.status = IntakeList.Status.SUBMITTED
                intake_list.submitted_at = timezone.now()
                # Clear the previous verdict so it is a fresh review.
                intake_list.pe_comments = ""
                intake_list.reviewed_at = None
            intake_list.save()

        return Response(
            IntakeListSerializer(intake_list, context=_serializer_context(request)).data
        )


class IntakeListReviewView(APIView):
    """
    PATCH /api/intake-lists/<id>/review/   (PE only)

    Body: {"action": "approve"|"request_changes"|"reject", "comments": "..."}
    """

    ACTION_STATUS = {
        "approve": IntakeList.Status.APPROVED,
        "request_changes": IntakeList.Status.REVISION_REQUESTED,
        "reject": IntakeList.Status.REJECTED,
    }

    def patch(self, request, pk):
        denied = _require_role(request, User.Role.PRINCIPAL_EMPLOYER)
        if denied:
            return denied

        intake_list = get_object_or_404(
            IntakeList.objects.select_related("project"), pk=pk
        )
        if intake_list.project.principal_employer_id != request.user.id:
            return Response(
                {"detail": "You can only review lists for your own projects."},
                status=status.HTTP_403_FORBIDDEN,
            )

        action = request.data.get("action")
        if action not in self.ACTION_STATUS:
            return Response(
                {"detail": f"action must be one of {list(self.ACTION_STATUS)}."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        intake_list.status = self.ACTION_STATUS[action]
        intake_list.pe_comments = request.data.get("comments", "")
        intake_list.reviewed_at = timezone.now()
        intake_list.save()
        return Response(
            IntakeListSerializer(intake_list, context=_serializer_context(request)).data
        )


# ---------------------------------------------------------------------------
# Gate security — REAL-TIME compliance, not an approval snapshot
# ---------------------------------------------------------------------------
class GateCheckView(APIView):
    """
    GET /api/gate-check/?aadhar=<number>   (Gate Security)

    Approval alone is no longer enough. The gate re-evaluates the worker's full
    compliance against the approved list's project **at scan time**, so a medical
    or PVC that lapsed after the PE signed off — or any regressed pillar — flips
    the decision to RED with the reason, instead of waving through a worker whose
    papers died last week.
    """

    def get(self, request):
        aadhar = (request.query_params.get("aadhar") or "").strip()
        if not aadhar:
            return Response(
                {"detail": "Provide ?aadhar=<number>."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        worker = (
            _prefetched(Worker.objects.filter(aadhar_number=aadhar)).first()
        )
        if worker is None:
            return Response(
                {
                    "access": "DENIED",
                    "reason_code": "UNKNOWN_WORKER",
                    "reason": "No worker found for this Aadhar number.",
                    "worker": None,
                    "checked_at": timezone.now().isoformat(),
                }
            )

        decision = worker.gate_decision()
        return Response(
            {
                **decision,
                "worker": {
                    "id": worker.id,
                    "name": worker.name,
                    "skill_type": worker.skill_type,
                    "aadhar_number": worker.aadhar_number,
                },
                "checked_at": timezone.now().isoformat(),
            }
        )


# ---------------------------------------------------------------------------
# Contractor Intake Workbench — mock OCR, strict verification, video heartbeat
# ---------------------------------------------------------------------------
def _check_not_expired(date_str, field_label, today):
    """Return ``(date, error_response)``. Rejects dates more than a year old.

    This is the strict boundary the whole platform hangs on: a document dated
    more than ``INTAKE_EXPIRY_DAYS`` ago is already expired on arrival and is
    never accepted, so an expired scan can't be verified into the system.
    """
    if not date_str:
        return None, Response(
            {"detail": f"{field_label} is required."},
            status=status.HTTP_400_BAD_REQUEST,
        )
    parsed = _parse_iso_date(date_str)
    if parsed is None:
        return None, Response(
            {"detail": f"{field_label} is not a valid ISO date."},
            status=status.HTTP_400_BAD_REQUEST,
        )
    if (today - parsed).days > INTAKE_EXPIRY_DAYS:
        return None, Response(
            {
                "detail": f"Rejected: {field_label} ({parsed.isoformat()}) is more "
                f"than {INTAKE_EXPIRY_DAYS} days old — the document is already "
                f"expired.",
                "expired": True,
            },
            status=status.HTTP_400_BAD_REQUEST,
        )
    return parsed, None


class VerifyDocumentView(APIView):
    """
    POST /api/intake/verify-document/   (Contractor)

    Strict verification-saving endpoint. For MEDICAL / POLICE it enforces that
    the exam/issue date is not older than 1 year and writes ``expiry_date``
    exactly 365 days out (the model does this on save). For IDENTITY it marks
    the named WorkerDocument as Verified.

    Accepts JSON or multipart/form-data. When multipart, the scanned document is
    read from the ``file`` field and pushed to the private bucket — so the
    contractor can upload the physical document and verify it on the spot.

    Body: {"worker": <id>, "doc_type": "MEDICAL"|"POLICE"|"IDENTITY", ...fields}
    """

    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def post(self, request):
        denied = _require_role(request, User.Role.CONTRACTOR)
        if denied:
            return denied

        worker = _owned_worker(request, request.data.get("worker"))
        doc_type = (request.data.get("doc_type") or "").upper()
        upload = request.FILES.get("file")
        today = timezone.now().date()
        context = _serializer_context(request, sign=True)

        if doc_type == "MEDICAL":
            exam_date, err = _check_not_expired(
                request.data.get("exam_date"), "Medical exam date", today
            )
            if err:
                return err
            rec, _ = IntakeMedicalRecord.objects.update_or_create(
                worker=worker,
                exam_date=exam_date,
                defaults={
                    "color_blindness": _as_bool(request.data.get("color_blindness", False)),
                    "vision": request.data.get("vision", ""),
                    "vertigo": _as_bool(request.data.get("vertigo", False)),
                    "blood_type": request.data.get("blood_type", ""),
                },
            )  # expiry_date computed in model.save()
            if upload:
                rec.storage_key = _store_upload(upload, "intake_docs") or ""
                rec.save(update_fields=["storage_key"])
            return Response(
                IntakeMedicalRecordSerializer(rec, context=context).data,
                status=status.HTTP_201_CREATED,
            )

        if doc_type == "POLICE":
            issue_date, err = _check_not_expired(
                request.data.get("issue_date"), "Police verification issue date", today
            )
            if err:
                return err
            rec, _ = IntakePoliceVerification.objects.update_or_create(
                worker=worker,
                issue_date=issue_date,
                defaults={
                    "certificate_number": request.data.get("certificate_number", ""),
                    "verification_status": request.data.get(
                        "verification_status", WorkerDocument.Status.VERIFIED
                    ),
                },
            )
            if upload:
                rec.storage_key = _store_upload(upload, "intake_docs") or ""
                rec.save(update_fields=["storage_key"])
            return Response(
                IntakePoliceVerificationSerializer(rec, context=context).data,
                status=status.HTTP_201_CREATED,
            )

        if doc_type == "IDENTITY":
            requirement_name = request.data.get("requirement_name", "Aadhar")
            requirement = RequirementMaster.objects.filter(name=requirement_name).first()
            if requirement is None:
                return Response(
                    {"detail": f"No requirement named '{requirement_name}'."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            defaults = {
                "document_number": request.data.get("document_number", ""),
                "verification_status": WorkerDocument.Status.VERIFIED,
                "rejection_reason": "",
            }
            # Expirable identity docs (e.g. Safety Training) may carry an expiry.
            expiry = request.data.get("expiry_date")
            if expiry:
                defaults["expiry_date"] = expiry
            doc, _ = WorkerDocument.objects.update_or_create(
                worker=worker, requirement=requirement, defaults=defaults
            )
            if upload:
                old_key = doc.storage_key
                doc.storage_key = _store_upload(upload, "worker_docs") or ""
                doc.file_url = ""
                doc.save(update_fields=["storage_key", "file_url"])
                if old_key and old_key != doc.storage_key:
                    storage.delete(old_key)
            return Response(
                WorkerDocumentSerializer(doc, context=context).data,
                status=status.HTTP_201_CREATED,
            )

        return Response(
            {"detail": "doc_type must be MEDICAL, POLICE or IDENTITY."},
            status=status.HTTP_400_BAD_REQUEST,
        )


class OcrExtractView(APIView):
    """
    POST /api/intake/ocr-extract/   (Contractor)

    Runs OCR on the uploaded scan and returns best-effort form fields for the
    given doc_type (IDENTITY | MEDICAL | POLICE). The contractor reviews and
    corrects the values, then commits via /verify-document/. Provider is
    env-selected; a failure returns empty fields plus a note rather than an error.
    """

    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        denied = _require_role(request, User.Role.CONTRACTOR)
        if denied:
            return denied

        upload = request.FILES.get("file")
        if upload is None:
            return Response(
                {"detail": "No file provided under form field 'file'."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        doc_type = (request.data.get("doc_type") or "").upper()
        if doc_type not in {"IDENTITY", "MEDICAL", "POLICE", "BANK"}:
            return Response(
                {"detail": "doc_type must be IDENTITY, MEDICAL, POLICE or BANK."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        requirement_name = request.data.get("requirement_name", "")
        # Which intake slot this file was dropped into, so we can tell an
        # Aadhaar card from a PAN card (both are doc_type IDENTITY).
        slot = (request.data.get("slot") or _slot_for(doc_type, requirement_name)).lower()
        today = timezone.now().date()

        fields, provider, note, text = ocr.extract_fields(
            upload.read(), upload.name, upload.content_type,
            doc_type, requirement_name, today,
        )
        check = doctype.check_document(slot, text, fields)
        return Response(
            {
                "form_type": doc_type,
                "slot": slot,
                "fields": fields,
                "provider": provider,
                "note": note,
                "check": check.as_dict(),
            }
        )


# ---------------------------------------------------------------------------
# Resume scanning
# ---------------------------------------------------------------------------
def _read_pages(request) -> list[tuple[bytes, str]]:
    """Collect resume pages from ``resume`` / ``resume[]`` / ``file`` fields."""
    files = (
        request.FILES.getlist("resume")
        or request.FILES.getlist("resume[]")
        or request.FILES.getlist("file")
    )
    return [(f.read(), f.content_type) for f in files if f]


def _persist_candidate_profile(worker, extraction, resume_key, contractor):
    """Write the parsed resume into the encrypted + filterable schema.

    PII goes through ``set_pii`` (ciphertext + blind index); everything else
    lands in plaintext, indexable columns. Skills are normalised into the shared
    vocabulary so two workers who both wrote "Welder" join the same row.
    """
    profile, _ = CandidateProfile.objects.get_or_create(
        worker=worker, defaults={"contractor": contractor}
    )
    profile.contractor = contractor
    profile.set_pii(
        name=extraction.name or "",
        phone=extraction.phone or "",
        email=extraction.email or "",
    )
    profile.place = extraction.place or ""
    profile.stream = extraction.stream or ""
    profile.category = extraction.category or ""
    profile.years_of_experience = extraction.years_of_experience
    profile.qualification = extraction.qualification or ""
    if resume_key:
        old_key = profile.resume_key
        profile.resume_key = resume_key
        if old_key and old_key != resume_key:
            storage.delete(old_key)
    profile.parser_provider = extraction.provider or ""
    profile.parse_note = (extraction.note or "")[:300]
    profile.save()

    profile.sync_name_tokens(extraction.name)

    # Replace the skill set wholesale — a re-scan is the new truth.
    profile.candidate_skills.all().delete()
    for raw in extraction.skills:
        skill = Skill.get_or_create_normalised(raw)
        if skill is not None:
            CandidateSkill.objects.get_or_create(profile=profile, skill=skill)
    return profile


class ResumeParseView(APIView):
    """
    POST /api/resume/parse/   (Contractor)

    Parse a resume and return the structured fields **without committing**, so
    the contractor can review the extraction before it is saved. Send
    ``worker=<id>`` as well to store it against that worker in the same call.
    """

    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        denied = _require_role(request, User.Role.CONTRACTOR)
        if denied:
            return denied

        pages = _read_pages(request)
        if not pages:
            return Response(
                {"detail": "No resume file provided under form field 'resume'."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        extraction = resume_parser.parse_resume(pages)

        worker_id = request.data.get("worker")
        if not worker_id:
            # Preview only — nothing is persisted, so nothing is encrypted yet.
            return Response({"committed": False, **extraction.as_dict()})

        worker = _owned_worker(request, worker_id)
        resume_key = storage.upload(
            pages[0][0], pages[0][1], "resumes", "resume"
        )
        profile = _persist_candidate_profile(worker, extraction, resume_key, request.user)
        return Response(
            {
                "committed": True,
                **extraction.as_dict(),
                "profile": CandidateProfileSerializer(
                    profile, context=_serializer_context(request, sign=True)
                ).data,
            },
            status=status.HTTP_201_CREATED,
        )


class CandidateSearchView(APIView):
    """
    GET /api/candidates/search/   (Contractor)

    Multi-attribute fuzzy filter over the contractor's own candidate profiles.

    Query params: ``q`` (name — matched through the blind index, never by
    decrypting), ``place``, ``stream``, ``category``, ``qualification``,
    ``skill`` (repeatable), ``min_experience``, ``max_experience``.
    """

    def get(self, request):
        denied = _require_role(request, User.Role.CONTRACTOR)
        if denied:
            return denied

        qs = (
            CandidateProfile.objects.filter(contractor=request.user)
            .select_related("worker")
            .prefetch_related("candidate_skills__skill")
        )

        params = request.query_params
        for field in ("place", "stream", "category", "qualification"):
            value = (params.get(field) or "").strip()
            if value:
                qs = qs.filter(**{f"{field}__icontains": value})

        for skill in [s for s in params.getlist("skill") if s.strip()]:
            qs = qs.filter(candidate_skills__skill__name__icontains=skill.strip().lower())

        for param, lookup in (("min_experience", "gte"), ("max_experience", "lte")):
            raw = params.get(param)
            if raw not in (None, ""):
                try:
                    qs = qs.filter(**{f"years_of_experience__{lookup}": int(raw)})
                except ValueError:
                    pass

        # Name search rides the blind index: hash the query the same way the
        # tokens were hashed and do an indexed equality probe. No decryption,
        # and the pepper never leaves this process.
        query = (params.get("q") or "").strip()
        if query:
            digests = crypto.name_query_tokens(query)
            if digests:
                qs = qs.filter(name_tokens__token_hash__in=digests)
            else:
                qs = qs.none()

        qs = qs.distinct()[:100]
        context = _serializer_context(request, sign=True)
        return Response(
            {
                "count": len(qs),
                "results": [
                    {
                        "profile": CandidateProfileSerializer(p, context=context).data,
                        "worker": {
                            "id": p.worker_id,
                            "name": p.worker.name,
                            "skill_type": p.worker.skill_type,
                            "aadhar_number": p.worker.aadhar_number,
                        },
                    }
                    for p in qs
                ],
            }
        )


# ---------------------------------------------------------------------------
# Unified worker onboarding — all 5 validation pillars + resume, in one pass
# ---------------------------------------------------------------------------
class UnifiedIntakeView(APIView):
    """
    POST /api/intake/onboard-worker/   (Contractor, multipart)

    Creates a worker and every document in a single submission:

      identity   name, aadhar_number, skill_type
      Aadhaar    aadhaar_file
      PAN        pan_file, pan_number
      Safety     safety_file, safety_expiry
      Medical    medical_file, exam_date, vision, blood_type,
                 color_blindness, vertigo
      PVC        pvc_file, certificate_number, issue_date
      Resume     resume_file  (parsed → encrypted PII + filterable profile)

    Order matters: **validate everything first**, then upload all files
    concurrently, then write the database rows in one transaction. Validating
    up front means an expired medical never leaves a half-created worker behind,
    and the concurrent upload keeps a six-document intake at roughly the cost of
    its slowest single file.
    """

    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        denied = _require_role(request, User.Role.CONTRACTOR)
        if denied:
            return denied

        data = request.data
        today = timezone.now().date()

        # -- 1. validate the worker identity --------------------------------
        name = (data.get("name") or "").strip()
        aadhar = (data.get("aadhar_number") or "").strip()
        skill = (data.get("skill_type") or "").strip()
        if not (name and aadhar and skill):
            return Response(
                {"detail": "name, aadhar_number and skill_type are required."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if len(aadhar) != 12 or not aadhar.isdigit():
            return Response(
                {"detail": "Aadhar number must be exactly 12 digits."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if Worker.objects.filter(aadhar_number=aadhar).exists():
            return Response(
                {"detail": f"A worker with Aadhar {aadhar} already exists."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # The Aadhaar card is the one document that must be on file — it is the
        # identity the gate scans against. Everything else can be typed in and
        # the scan supplied later.
        if request.FILES.get("aadhaar_file") is None:
            return Response(
                {
                    "detail": "An Aadhaar card scan is required. Every other document "
                    "is optional and its details can be entered directly.",
                    "missing_document": "aadhaar_file",
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        # -- 2. validate the dated pillars before touching anything ---------
        exam_date = issue_date = None
        if request.FILES.get("medical_file") or data.get("exam_date"):
            exam_date, err = _check_not_expired(
                data.get("exam_date"), "Medical exam date", today
            )
            if err:
                return err
        if request.FILES.get("pvc_file") or data.get("issue_date"):
            issue_date, err = _check_not_expired(
                data.get("issue_date"), "Police verification issue date", today
            )
            if err:
                return err

        requirements_by_name = {r.name: r for r in RequirementMaster.objects.all()}

        # -- 3. upload every attached file concurrently ---------------------
        slots = [
            ("aadhaar", "aadhaar_file", "worker_docs"),
            ("pan", "pan_file", "worker_docs"),
            ("safety", "safety_file", "worker_docs"),
            ("medical", "medical_file", "intake_docs"),
            ("pvc", "pvc_file", "intake_docs"),
            ("bank", "bank_file", "intake_docs"),
            ("resume", "resume_file", "resumes"),
        ]
        items, resume_pages = [], []
        for slot, field_name, prefix in slots:
            upload = request.FILES.get(field_name)
            if upload is None:
                continue
            blob = upload.read()
            if not blob:
                continue
            items.append({
                "slot": slot,
                "data": blob,
                "content_type": upload.content_type,
                "prefix": prefix,
                "filename": upload.name,
            })
            if slot == "resume":
                resume_pages.append((blob, upload.content_type))

        # -- 3b. verify each scan really is the document its slot expects ---
        # Which slots to check server-side is a cost decision: each check is an
        # OCR round trip. The default verifies the Aadhaar only — it is the
        # mandatory, identity-critical one — while the browser has already
        # checked every slot as the file was attached. Set
        # VERIFY_DOCUMENT_TYPES="all" to enforce them all, or "none" to skip.
        policy = getattr(settings, "VERIFY_DOCUMENT_TYPES", "aadhaar").lower()
        if policy != "none":
            to_check = [i for i in items if i["slot"] != "resume"]
            if policy != "all":
                to_check = [i for i in to_check if i["slot"] == "aadhaar"]
            for item in to_check:
                fields, _provider, _note, text = ocr.extract_fields(
                    item["data"], item["filename"], item["content_type"],
                    _doc_type_for_slot(item["slot"]), _requirement_for_slot(item["slot"]),
                    today,
                )
                verdict = doctype.check_document(item["slot"], text, fields)
                if not verdict.ok:
                    return Response(
                        {"detail": verdict.message, "document_check": verdict.as_dict(),
                         "slot": item["slot"]},
                        status=status.HTTP_400_BAD_REQUEST,
                    )

        keys = storage.upload_many(items)

        # -- 4. parse the resume (never fatal) ------------------------------
        extraction = (
            resume_parser.parse_resume(resume_pages) if resume_pages else None
        )

        # -- 5. one transaction for the whole worker ------------------------
        created_docs = []
        try:
            with transaction.atomic():
                worker = Worker.objects.create(
                    name=name,
                    aadhar_number=aadhar,
                    skill_type=skill,
                    contractor=request.user,
                )

                def identity_doc(requirement_name, slot, number="", expiry=None):
                    requirement = requirements_by_name.get(requirement_name)
                    if requirement is None or (slot not in keys and not number):
                        return
                    doc = WorkerDocument.objects.create(
                        worker=worker,
                        requirement=requirement,
                        document_number=number or "",
                        storage_key=keys.get(slot, ""),
                        expiry_date=expiry,
                        verification_status=WorkerDocument.Status.VERIFIED,
                    )
                    created_docs.append(doc)

                identity_doc("Aadhar", "aadhaar", aadhar)
                identity_doc("PAN", "pan", (data.get("pan_number") or "").strip())
                identity_doc(
                    "Safety Training",
                    "safety",
                    (data.get("safety_number") or "").strip(),
                    _parse_iso_date(data.get("safety_expiry")),
                )

                if exam_date is not None:
                    IntakeMedicalRecord.objects.create(
                        worker=worker,
                        exam_date=exam_date,
                        color_blindness=_as_bool(data.get("color_blindness", False)),
                        vertigo=_as_bool(data.get("vertigo", False)),
                        vision=data.get("vision", ""),
                        blood_type=data.get("blood_type", ""),
                        storage_key=keys.get("medical", ""),
                    )

                if issue_date is not None:
                    IntakePoliceVerification.objects.create(
                        worker=worker,
                        issue_date=issue_date,
                        certificate_number=data.get("certificate_number", ""),
                        verification_status=data.get(
                            "verification_status", WorkerDocument.Status.VERIFIED
                        ),
                        storage_key=keys.get("pvc", ""),
                    )

                account_number = (data.get("bank_account_number") or "").strip()
                if account_number or keys.get("bank"):
                    bank = WorkerBankAccount(
                        worker=worker,
                        ifsc=(data.get("ifsc") or "").strip().upper(),
                        bank_name=(data.get("bank_name") or "").strip(),
                        account_holder_name=(data.get("account_holder_name") or "").strip(),
                        storage_key=keys.get("bank", ""),
                    )
                    bank.set_account_number(account_number)
                    bank.save()

                if extraction is not None:
                    _persist_candidate_profile(
                        worker, extraction, keys.get("resume", ""), request.user
                    )
        except Exception:
            # The DB write failed after the objects landed in the bucket —
            # clean them up rather than leaving orphans behind.
            for key in keys.values():
                storage.delete(key)
            raise

        worker = _prefetched(Worker.objects.filter(pk=worker.pk)).first()
        context = _serializer_context(request, sign=True)
        return Response(
            {
                "worker": WorkerSerializer(worker, context=context).data,
                "compliance": worker.compliance_snapshot(),
                "documents_stored": sorted(keys.keys()),
                "resume": extraction.as_dict() if extraction else None,
            },
            status=status.HTTP_201_CREATED,
        )


# ---------------------------------------------------------------------------
# Storage — fresh signed links + the local-fallback object server
# ---------------------------------------------------------------------------
class SignedUrlView(APIView):
    """
    POST /api/storage/signed-url/   (Contractor / PE)

    Body: ``{"keys": ["worker_docs/2026/…", …]}``
    Returns ``{key: url}`` with links valid for ``PRESIGN_EXPIRY_SECONDS``.

    Batching matters: a document table needs one link per row, and one HTTP
    round trip per row would make the table crawl.
    """

    def post(self, request):
        denied = _require_role(
            request, User.Role.CONTRACTOR, User.Role.PRINCIPAL_EMPLOYER
        )
        if denied:
            return denied

        keys = request.data.get("keys") or []
        if not isinstance(keys, list):
            return Response(
                {"detail": "keys must be a list of object keys."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        return Response({"urls": storage.signed_urls(keys[:200])})


class StorageObjectView(APIView):
    """
    GET /api/storage/object/?key=…&expires=…&signature=…

    Serves an object from the **local fallback** backend after verifying the
    HMAC signature and expiry. Unauthenticated by design: the signed URL *is*
    the capability, exactly as with a Supabase signed link. When Supabase is
    configured this route is never used — links point at Supabase directly.
    """

    permission_classes = [AllowAny]

    def get(self, request):
        backend = storage.get_storage()
        if not isinstance(backend, storage.LocalSignedStorage):
            raise Http404("Objects are served directly by Supabase.")

        key = request.query_params.get("key") or ""
        signature = request.query_params.get("signature") or ""
        try:
            expires = int(request.query_params.get("expires") or 0)
        except ValueError:
            expires = 0

        if not backend.verify(key, expires, signature):
            return Response(
                {"detail": "This link is invalid or has expired."},
                status=status.HTTP_403_FORBIDDEN,
            )

        try:
            path = backend.path_for(key)
        except storage.StorageError:
            raise Http404("Unknown object.")
        if not path.exists():
            raise Http404("Unknown object.")
        return FileResponse(path.open("rb"))


# ---------------------------------------------------------------------------
# Trade test — Contractor-administered practical MCQ exam
# ---------------------------------------------------------------------------
def _trade_test_state(worker):
    """Current attempt bookkeeping for a worker."""
    attempts_used = worker.trade_test_attempts.count()
    return {
        "attempts_used": attempts_used,
        "attempts_remaining": max(0, TRADE_TEST_MAX_ATTEMPTS - attempts_used),
        "status": worker.trade_test_status,
        "passed": worker.trade_test_status == Worker.TradeTestStatus.PASSED,
        "locked": worker.trade_test_status == Worker.TradeTestStatus.FAILED,
    }


class TradeTestStartView(APIView):
    """
    GET /api/trade-test/start/?worker_id=<id>   (Contractor)

    Validates the worker has remaining attempts and has not already passed, then
    returns exactly 5 random questions for the worker's skill category — WITHOUT
    the correct answers.
    """

    def get(self, request):
        denied = _require_role(request, User.Role.CONTRACTOR)
        if denied:
            return denied

        worker = _owned_worker(request, request.query_params.get("worker_id"))
        state = _trade_test_state(worker)

        if state["passed"]:
            return Response(
                {"detail": "This worker has already passed the trade test.", **state},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if state["locked"] or state["attempts_remaining"] <= 0:
            return Response(
                {"detail": "No attempts remaining — profile is locked as Failed.", **state},
                status=status.HTTP_400_BAD_REQUEST,
            )

        category = category_for_skill(worker.skill_type)
        questions = list(
            TradeTestQuestion.objects.filter(skill_type=category).order_by("?")[
                :TRADE_TEST_QUESTION_COUNT
            ]
        )
        if len(questions) < TRADE_TEST_QUESTION_COUNT:
            return Response(
                {
                    "detail": f"Only {len(questions)} questions exist for {category}; "
                    f"need {TRADE_TEST_QUESTION_COUNT}. Seed more questions."
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        return Response(
            {
                "worker_id": worker.id,
                "worker_name": worker.name,
                "category": category,
                "attempt_number": state["attempts_used"] + 1,
                "attempts_remaining": state["attempts_remaining"],
                "pass_mark": TRADE_TEST_PASS_MARK,
                "questions": TradeTestQuestionSerializer(questions, many=True).data,
            }
        )


class TradeTestSubmitView(APIView):
    """
    POST /api/trade-test/submit-attempt/   (Contractor)

    Body: {"worker_id": <id>, "answers": [{"question_id": <id>, "selected_option": "A"}]}

    Scores server-side, records the attempt, and updates the worker's trade-test
    status: PASSED at >= 3/5, or FAILED (locked) once the 3rd attempt is used up.
    """

    def post(self, request):
        denied = _require_role(request, User.Role.CONTRACTOR)
        if denied:
            return denied

        worker = _owned_worker(request, request.data.get("worker_id"))
        state = _trade_test_state(worker)
        if state["passed"]:
            return Response(
                {"detail": "Worker has already passed.", **state},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if state["locked"] or state["attempts_remaining"] <= 0:
            return Response(
                {"detail": "No attempts remaining — profile is locked.", **state},
                status=status.HTTP_400_BAD_REQUEST,
            )

        answers = request.data.get("answers") or []
        if not isinstance(answers, list) or not answers:
            return Response(
                {"detail": "answers must be a non-empty list."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Score server-side against the stored correct_option (never trust client).
        selected = {}
        for a in answers:
            try:
                selected[int(a.get("question_id"))] = (a.get("selected_option") or "").upper()
            except (TypeError, ValueError):
                continue

        questions = TradeTestQuestion.objects.filter(id__in=selected.keys())
        score = sum(1 for q in questions if selected.get(q.id) == q.correct_option)
        passed = score >= TRADE_TEST_PASS_MARK
        attempt_number = state["attempts_used"] + 1

        with transaction.atomic():
            TradeTestAttempt.objects.create(
                worker=worker,
                attempt_number=attempt_number,
                score=score,
                is_passed=passed,
            )
            if passed:
                worker.trade_test_status = Worker.TradeTestStatus.PASSED
            elif attempt_number >= TRADE_TEST_MAX_ATTEMPTS:
                worker.trade_test_status = Worker.TradeTestStatus.FAILED
            worker.save(update_fields=["trade_test_status"])

        attempts_remaining = max(0, TRADE_TEST_MAX_ATTEMPTS - attempt_number)
        return Response(
            {
                "worker_id": worker.id,
                "score": score,
                "total": TRADE_TEST_QUESTION_COUNT,
                "pass_mark": TRADE_TEST_PASS_MARK,
                "is_passed": passed,
                "attempt_number": attempt_number,
                "attempts_remaining": attempts_remaining,
                "trade_test_status": worker.trade_test_status,
                "locked": worker.trade_test_status == Worker.TradeTestStatus.FAILED,
            },
            status=status.HTTP_201_CREATED,
        )


# ---------------------------------------------------------------------------
# Safety Training video — watch-progress heartbeat
# ---------------------------------------------------------------------------
class SafetyVideoHeartbeatView(APIView):
    """
    POST /api/safety-video/heartbeat/   (Contractor)

    Records how much of the mandatory safety induction video a worker has
    watched. ``is_completed`` flips to True at 100%. Progress never moves
    backwards.

    Body: {"worker": <id>, "progress_percentage": <0-100>}
    """

    def post(self, request):
        denied = _require_role(request, User.Role.CONTRACTOR)
        if denied:
            return denied

        worker = _owned_worker(request, request.data.get("worker"))
        try:
            pct = int(request.data.get("progress_percentage", 0))
        except (TypeError, ValueError):
            return Response(
                {"detail": "progress_percentage must be an integer."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        pct = max(0, min(100, pct))

        sv, _ = SafetyTrainingProgress.objects.get_or_create(worker=worker)
        # Monotonic: never regress a previously higher watermark.
        sv.progress_percentage = max(sv.progress_percentage, pct)
        sv.is_completed = sv.progress_percentage >= 100
        sv.save()
        return Response(
            {
                "worker": worker.id,
                "progress_percentage": sv.progress_percentage,
                "is_completed": sv.is_completed,
            }
        )
